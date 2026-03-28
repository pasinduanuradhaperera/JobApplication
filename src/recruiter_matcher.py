"""
recruiter_matcher.py — Finds best-fit jobs for a CV and exports ranked results to Excel.
"""

import json
import math
import os
import re
from datetime import datetime

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from job_search import search_jobs

DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEEPSEEK_MODEL = "deepseek-chat"


def _get_client() -> OpenAI:
    return OpenAI(api_key=os.environ["DEEPSEEK_API_KEY"], base_url=DEEPSEEK_BASE_URL)


def _extract_first_json(raw: str) -> str:
    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("No JSON object found in model response.")
    return raw[start : end + 1]


def suggest_job_queries_from_cv(cv_text: str, max_queries: int = 5) -> list[str]:
    """Use AI to suggest role-based job search queries from CV content."""
    prompt = f"""
You are a senior recruiter. Based on this CV, suggest the best job-search queries.

CV:
{cv_text[:6000]}

Return valid JSON only with this exact schema:
{{
  "queries": ["query1", "query2", "query3"]
}}

Rules:
- 3 to {max_queries} queries.
- Keep each query concise (2 to 6 words).
- Focus on realistic role titles and skill combinations.
"""
    raw = _get_client().chat.completions.create(
        model=DEEPSEEK_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
    ).choices[0].message.content.strip()

    data = json.loads(_extract_first_json(raw))
    queries = [str(q).strip() for q in data.get("queries", []) if str(q).strip()]
    if not queries:
        raise ValueError("Model returned no queries.")
    return queries[:max_queries]


def _fallback_queries(cv_text: str) -> list[str]:
    text = cv_text.lower()
    candidates = [
        ("Python Developer", ["python", "django", "flask", "fastapi"]),
        ("Data Analyst", ["sql", "tableau", "power bi", "analytics"]),
        ("Machine Learning Engineer", ["machine learning", "tensorflow", "pytorch", "ml"]),
        ("Backend Engineer", ["api", "backend", "microservices", "server"]),
        ("Full Stack Developer", ["react", "node", "javascript", "full stack"]),
    ]
    scored = []
    for title, kws in candidates:
        score = sum(1 for kw in kws if kw in text)
        scored.append((score, title))
    scored.sort(reverse=True)
    picks = [title for score, title in scored if score > 0]
    return picks[:5] or ["Python Developer", "Software Engineer", "Backend Engineer"]


def _dedupe_jobs(jobs: list[dict]) -> list[dict]:
    seen = set()
    unique = []
    for job in jobs:
        key = (
            (job.get("title") or "").strip().lower(),
            (job.get("company") or "").strip().lower(),
            (job.get("location") or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(job)
    return unique


def _fallback_rank(cv_text: str, jobs: list[dict], top_n: int) -> list[dict]:
    tokens = set(re.findall(r"[a-zA-Z]{3,}", cv_text.lower()))
    ranked = []
    for job in jobs:
        hay = f"{job.get('title', '')} {job.get('snippet', '')}".lower()
        overlap = sum(1 for t in tokens if t in hay)
        score = min(95, 40 + overlap)
        ranked.append(
            {
                "score": score,
                "reason": "Keyword overlap with CV profile.",
                **job,
            }
        )
    ranked.sort(key=lambda x: x["score"], reverse=True)
    return ranked[:top_n]


def rank_jobs_for_cv(cv_text: str, jobs: list[dict], top_n: int = 25) -> list[dict]:
    """Score and rank jobs against CV as a recruiter would."""
    if not jobs:
        return []

    payload_lines = []
    for i, job in enumerate(jobs, start=1):
        payload_lines.append(
            f"{i}. TITLE: {job.get('title','')} | COMPANY: {job.get('company','')} | "
            f"LOCATION: {job.get('location','')} | SOURCE: {job.get('source','')} | "
            f"DESCRIPTION: {job.get('snippet','')[:300]}"
        )

    prompt = f"""
        I will upload my resume. Act as an Al recruiter and Job hunting machine. Analyze my resume in depth to identify the most suitable fresher or entry-level roles or suitable level I should target in Sri Lanka.onsite,hybrid,online or whatever.even another country remote job it okay.. Find real companies currently hiring across startups, scale-ups, MNCs, consulting firms, and both tech and non-tech sectors, and provide verified application links for each opportunity.
        Match every job with my profile and give a fit score out of 100. Create a prioritized job application list categorized into high-probability, medium-probability, and stretch roles. Curate a list of jobs with application links.

Candidate CV:
{cv_text[:7000]}

Jobs:
{chr(10).join(payload_lines)}

Return valid JSON only with this exact schema:
{{
  "ranked": [
    {{"index": 1, "score": 88, "reason": "Short reason"}}
  ]
}}

Rules:
- Include up to {top_n} items.
- Score must be 0 to 100.
- index is the 1-based number from the jobs list.
- reason must be concise (max 18 words).
- Prioritize skills fit, seniority fit, and role alignment.
"""

    try:
        raw = _get_client().chat.completions.create(
            model=DEEPSEEK_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        ).choices[0].message.content.strip()
        data = json.loads(_extract_first_json(raw))
        ranked = []
        for item in data.get("ranked", []):
            idx = int(item.get("index", 0))
            if not (1 <= idx <= len(jobs)):
                continue
            score = max(0, min(100, int(item.get("score", 0))))
            reason = str(item.get("reason", "")).strip() or "Strong profile alignment."
            ranked.append({"score": score, "reason": reason, **jobs[idx - 1]})

        ranked.sort(key=lambda x: x["score"], reverse=True)
        if ranked:
            return ranked[:top_n]
    except Exception:
        pass

    return _fallback_rank(cv_text, jobs, top_n)


def find_best_jobs_for_cv(
    cv_text: str,
    location: str = "",
    max_results: int = 25,
) -> tuple[list[dict], list[str]]:
    """Generate role queries from CV, collect jobs, and return ranked best matches."""
    if not str(cv_text or "").strip():
        raise ValueError("CV text is empty. Please provide a valid CV PDF.")

    try:
        max_results = int(max_results)
    except (TypeError, ValueError):
        max_results = 25
    max_results = max(1, min(100, max_results))

    try:
        queries = suggest_job_queries_from_cv(cv_text)
    except Exception:
        queries = _fallback_queries(cv_text)

    per_query = max(8, math.ceil((max_results * 2) / max(len(queries), 1)))
    all_jobs: list[dict] = []
    for query in queries:
        try:
            all_jobs.extend(search_jobs(query, location, per_query))
        except Exception:
            continue

    unique_jobs = _dedupe_jobs(all_jobs)
    ranked = rank_jobs_for_cv(cv_text, unique_jobs, top_n=max_results)
    return ranked, queries


def _auto_column_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 70)


def save_ranked_jobs_to_excel(ranked_jobs: list[dict], output_path: str) -> str:
    """Write ranked jobs into a new Excel workbook."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Best Matches"
    headers = [
        "Date", "Rank", "Suitability Score", "Job Title", "Company", "Location",
        "Job URL", "Source", "Min Experience", "Reason",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for rank, job in enumerate(ranked_jobs, start=1):
        ws.append(
            [
                now,
                rank,
                job.get("score", ""),
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("url", ""),
                job.get("source", ""),
                job.get("exp", ""),
                job.get("reason", ""),
            ]
        )

    _auto_column_width(ws)
    wb.save(output_path)
    return output_path
