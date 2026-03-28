"""
tracker.py — Logs every application and job search result to Excel workbooks.

Files created (inside output/):
  - applications.xlsx   : one row per application (Modules 1 & 2)
  - job_search.xlsx     : one row per job found   (Module 3)
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")

APP_FILE  = os.path.join(OUTPUT_DIR, "applications.xlsx")
JOBS_FILE = os.path.join(OUTPUT_DIR, "job_search.xlsx")

APP_HEADERS = [
    "Date", "Company", "Role", "Recipient Email",
    "Cover Letter PDF", "Email Subject", "Email Body",
    "Status",           # "Draft" or "Sent"
    "Notes",
]

JOBS_HEADERS = [
    "Date Found", "Job Title", "Company", "Location",
    "Job URL", "Description Snippet", "Source",
]


# ─── helpers ──────────────────────────────────────────────────────────────────

def _ensure_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _style_header_row(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_column_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _open_or_create(filepath: str, headers: list) -> tuple:
    """Return (workbook, worksheet). Creates file with headers if it doesn't exist."""
    _ensure_dir()
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        _style_header_row(ws)
    return wb, ws


# ─── public API ───────────────────────────────────────────────────────────────

def log_application(
    company: str,
    role: str,
    recipient_email: str,
    cover_letter_pdf: str,
    email_subject: str,
    email_body: str,
    status: str = "Draft",
    notes: str = "",
) -> None:
    """Append one row to applications.xlsx."""
    wb, ws = _open_or_create(APP_FILE, APP_HEADERS)
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        company,
        role,
        recipient_email,
        cover_letter_pdf,
        email_subject,
        email_body,
        status,
        notes,
    ])
    _auto_column_width(ws)
    wb.save(APP_FILE)
    print(f"[tracker] Application logged → {APP_FILE}")


def update_application_status(company: str, role: str, new_status: str) -> None:
    """Update the Status column for the most recent matching row."""
    if not os.path.exists(APP_FILE):
        return
    wb = load_workbook(APP_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    status_col = headers.index("Status") + 1
    company_col = headers.index("Company") + 1
    role_col = headers.index("Role") + 1

    for row in reversed(list(ws.iter_rows(min_row=2))):
        if row[company_col - 1].value == company and row[role_col - 1].value == role:
            row[status_col - 1].value = new_status
            break
    wb.save(APP_FILE)


def log_job(
    title: str,
    company: str,
    location: str,
    url: str,
    snippet: str,
    source: str = "Web Search",
) -> None:
    """Append one row to job_search.xlsx."""
    wb, ws = _open_or_create(JOBS_FILE, JOBS_HEADERS)
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        title,
        company,
        location,
        url,
        snippet,
        source,
    ])
    _auto_column_width(ws)
    wb.save(JOBS_FILE)
