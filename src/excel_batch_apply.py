"""
excel_batch_apply.py — Batch application flow driven by an Excel sheet.

Expected columns in Excel (header row, case-insensitive):
  - company
  - role
  - recipient_email
  - cv_path
Optional:
  - extras
"""

import os
import re
from datetime import datetime
from typing import Callable

from openpyxl import load_workbook

from cv_parser import extract_cv_text
from generator import (
    generate_cover_letter, generate_email, save_cover_letter_pdf,
    regenerate_cover_letter, regenerate_email,
)
from email_sender import send_application_email
from tracker import log_application

REQUIRED_COLUMNS = {"company", "role", "recipient_email", "cv_path"}
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
HEADER_ALIASES = {
    "company": {
        "company", "company_name", "organization", "employer",
    },
    "role": {
        "role", "job_title", "title", "position", "designation",
    },
    "recipient_email": {
        "recipient_email", "email", "hr_email", "hiring_manager_email", "recipient",
    },
    "cv_path": {
        "cv_path", "resume_path", "cv", "resume", "cv_file", "resume_file",
    },
    "extras": {
        "extras", "extra", "notes", "additional_info", "additional_details",
    },
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return text.replace(" ", "_")


def _canonical_header(header: str) -> str:
    """Map known header aliases to canonical names used by the workflow."""
    for canonical, aliases in HEADER_ALIASES.items():
        if header in aliases:
            return canonical
    return header


def _is_valid_email(email: str) -> bool:
    return bool(EMAIL_RE.match((email or "").strip()))


def load_batch_rows(excel_path: str) -> list[dict]:
    """Read and validate batch rows from Excel."""
    wb = load_workbook(excel_path)
    ws = wb.active

    if ws.max_row < 1:
        raise ValueError(
            "Excel sheet appears to be empty. Add a header row first.\n"
            "Required columns: company, role, recipient_email, cv_path"
        )

    raw_headers = [_normalize_header(cell.value) for cell in ws[1]]
    headers = [_canonical_header(h) for h in raw_headers]
    missing = sorted(REQUIRED_COLUMNS - set(headers))
    if missing:
        found = [h for h in headers if h]
        found_text = ", ".join(found) if found else "(no header names detected)"
        raise ValueError(
            "Excel header format is not recognized.\n"
            "Missing required columns: "
            + ", ".join(missing)
            + "\nDetected columns: "
            + found_text
            + "\nUse these exact headers (recommended): company, role, recipient_email, cv_path"
            + "\nAlso accepted aliases include: company_name, job_title, email, resume_path, notes"
        )

    rows: list[dict] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: row[i] for i in range(len(headers))}
        if not any(str(v or "").strip() for v in data.values()):
            continue
        data["_row_number"] = row_num
        rows.append(data)
    return rows


def _get_status_column_index(headers_raw: list) -> int | None:
    """Find the index of the 'status' column if it exists."""
    for i, h in enumerate(headers_raw):
        if _normalize_header(h) == "status":
            return i + 1
    return None


def _skip_already_applied(excel_path: str) -> tuple[list[dict], bool]:
    """Load rows and check for already-applied ones. Return (rows_to_process, has_applied)."""
    wb = load_workbook(excel_path)
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]
    status_col = _get_status_column_index(raw_headers)
    
    rows = load_batch_rows(excel_path)
    
    applied_rows = []
    pending_rows = []
    has_applied = False
    
    for row in rows:
        row_num = row.get("_row_number", 0)
        if status_col and row_num >= 2:
            status_val = ws.cell(row_num, status_col).value
            if status_val in ("Sent", "Draft", "Skipped"):
                applied_rows.append(row)
                has_applied = True
                continue
        pending_rows.append(row)
    
    return pending_rows, has_applied


def run_batch_apply_workflow(
    excel_path: str,
    output_dir: str,
    confirm_fn: Callable[[str], bool],
    preview_fn: Callable[[str, str], None],
    ask_fn: Callable[[str, str], str] | None = None,
) -> None:
    """Generate and review applications one by one from Excel rows before sending."""
    rows, has_applied = _skip_already_applied(excel_path)
    
    if not rows:
        if has_applied:
            print("  All rows in this Excel file have already been processed.")
            print("  To reprocess: add new rows, or manually clear the 'Status' column.")
        else:
            print("  No application rows found in the Excel sheet.")
        return
    
    if has_applied:
        print(f"  Found {len(rows)} new row(s) to process (skipping already-applied positions).")

    cv_cache: dict[str, str] = {}
    sent = 0
    drafted = 0
    skipped = 0

    wb_write = load_workbook(excel_path)
    ws_write = wb_write.active
    raw_headers_write = [cell.value for cell in ws_write[1]]
    status_col_write = _get_status_column_index(raw_headers_write)

    for idx, row in enumerate(rows, start=1):
        company = str(row.get("company", "") or "").strip()
        role = str(row.get("role", "") or "").strip()
        recipient = str(row.get("recipient_email", "") or "").strip()
        cv_path = str(row.get("cv_path", "") or "").strip().strip('"').strip("'")
        cv_path = os.path.abspath(os.path.expanduser(cv_path)) if cv_path else ""
        extras = str(row.get("extras", "") or "").strip()
        row_num = row.get("_row_number", "?")

        print("\n" + "=" * 68)
        print(f"  Processing {idx}/{len(rows)}  (Excel row {row_num})")
        print(f"  Company: {company or '—'}")
        print(f"  Role   : {role or '—'}")
        print(f"  Email  : {recipient or '—'}")

        if not (company and role and recipient and cv_path):
            print("  ✗ Missing required values (company/role/recipient_email/cv_path). Skipping row.")
            skipped += 1
            log_application(
                company=company,
                role=role,
                recipient_email=recipient,
                cover_letter_pdf="",
                email_subject="",
                email_body="",
                status="Skipped",
                notes=f"Excel row {row_num}: Missing required values.",
            )
            continue

        if not _is_valid_email(recipient):
            print(f"  ✗ Invalid recipient email: {recipient}")
            skipped += 1
            log_application(
                company=company,
                role=role,
                recipient_email=recipient,
                cover_letter_pdf="",
                email_subject="",
                email_body="",
                status="Skipped",
                notes=f"Excel row {row_num}: Invalid recipient email format.",
            )
            continue

        if not (os.path.isfile(cv_path) and cv_path.lower().endswith(".pdf")):
            print(f"  ✗ Invalid CV path: {cv_path}")
            skipped += 1
            log_application(
                company=company,
                role=role,
                recipient_email=recipient,
                cover_letter_pdf="",
                email_subject="",
                email_body="",
                status="Skipped",
                notes=f"Excel row {row_num}: CV path invalid or not a PDF.",
            )
            continue

        try:
            if cv_path not in cv_cache:
                cv_cache[cv_path] = extract_cv_text(cv_path)
            cv_text = cv_cache[cv_path]

            print("  Generating cover letter ...")
            cover_letter = generate_cover_letter(cv_text, company, role, extras)

            print("  Generating email ...")
            email_data = generate_email(cv_text, company, role, cover_letter)
        except Exception as e:
            print(f"  ✗ Generation failed: {e}")
            skipped += 1
            log_application(
                company=company,
                role=role,
                recipient_email=recipient,
                cover_letter_pdf="",
                email_subject="",
                email_body="",
                status="Skipped",
                notes=f"Excel row {row_num}: Generation failed ({e}).",
            )
            continue

        # ── Preview & Regeneration Loop ──
        approved = False
        while True:
            preview_fn("COVER LETTER", cover_letter)
            preview_fn(f"EMAIL SUBJECT: {email_data['subject']}\n\n  EMAIL BODY", email_data["body"])

            if confirm_fn("Are you happy with this draft?"):
                approved = True
                break

            if not confirm_fn("Would you like to regenerate it?"):
                break

            feedback = ""
            if confirm_fn("Provide suggestions to improve the regeneration?"):
                if ask_fn:
                    feedback = ask_fn("Your suggestions (or press Enter for generic regeneration)", "")
                else:
                    feedback = input("Your suggestions (or press Enter for generic regeneration): ").strip()

            print("  Regenerating cover letter ...")
            cover_letter = regenerate_cover_letter(cv_text, company, role, extras, feedback)
            print("  Regenerating email ...")
            email_data = regenerate_email(cv_text, company, role, cover_letter, feedback)
            print("  ✓ Regenerated.")

        if not approved:
            print("  Skipped by user.")
            skipped += 1
            log_application(
                company=company,
                role=role,
                recipient_email=recipient,
                cover_letter_pdf="",
                email_subject=email_data["subject"],
                email_body=email_data["body"],
                status="Skipped",
                notes=f"Excel row {row_num}: Rejected during review.",
            )
            continue

        safe_name = f"{company}_{role}_row{row_num}".replace(" ", "_").replace("/", "-")
        pdf_path = os.path.join(output_dir, f"CoverLetter_{safe_name}.pdf")
        save_cover_letter_pdf(cover_letter, pdf_path)
        print(f"  ✓ Cover letter saved → {pdf_path}")

        status = "Draft"
        notes = f"Excel row {row_num}: Approved by user."
        if confirm_fn(f"Send this email now to {recipient}?"):
            try:
                send_application_email(
                    to_address=recipient,
                    subject=email_data["subject"],
                    body=email_data["body"],
                    cover_letter_pdf_path=pdf_path,
                    cv_path=cv_path,
                )
                status = "Sent"
                notes = f"Excel row {row_num}: Sent successfully."
                sent += 1
                print("  ✓ Email sent.")
            except Exception as e:
                status = "Draft"
                notes = f"Excel row {row_num}: Send failed ({e})."
                drafted += 1
                print(f"  ✗ Send failed, kept as Draft: {e}")
        else:
            drafted += 1
            print("  Kept as draft.")

        log_application(
            company=company,
            role=role,
            recipient_email=recipient,
            cover_letter_pdf=pdf_path,
            email_subject=email_data["subject"],
            email_body=email_data["body"],
            status=status,
            notes=notes,
        )

        if status_col_write:
            ws_write.cell(row_num, status_col_write).value = status
            try:
                wb_write.save(excel_path)
            except Exception:
                pass

    print("\n" + "=" * 68)
    print("  Batch complete")
    print(f"  Sent   : {sent}")
    print(f"  Drafts : {drafted}")
    print(f"  Skipped: {skipped}")
    print("=" * 68)
